"""Generate fit-gap-workshop-tracker.xlsx with 4 sheets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

OUTPUT = r"C:\Users\dimos\Implementation\artefacts\03-fit-gap-register\fit-gap-workshop-tracker.xlsx"

# Styles
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

def style_body(ws, rows_start, rows_end, cols_end):
    for r in range(rows_start, rows_end + 1):
        for c in range(1, cols_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL

def add_dv(ws, options, col_letter, start_row=2, end_row=50):
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv)

wb = Workbook()

# ========== Sheet 1: Requirements ==========
ws1 = wb.active
ws1.title = "Requirements"
headers1 = ["Req ID", "Process Area", "Description", "Priority", "MoSCoW", "Pre-Score", "Pre-Score Notes"]
widths1 = [12, 20, 50, 12, 12, 16, 40]
style_header(ws1, headers1, widths1)

samples1 = [
    ["REQ-001", "Customer Onboarding", "Digital KYC verification with biometric identity checks", "High", "Must", "Full Fit", "T24 Digital KYC module supports biometric verification natively"],
    ["REQ-002", "Customer Onboarding", "Automated PEP and sanctions screening during onboarding", "High", "Must", "Full Fit", "Compliance screening module includes PEP/sanctions lists"],
    ["REQ-003", "Account Management", "Multi-currency account opening with real-time FX conversion", "Medium", "Should", "Partial Fit", "Multi-currency supported; real-time FX requires integration"],
    ["REQ-008", "Lending", "Dynamic interest rate adjustment based on risk profile changes", "Medium", "Should", "No Fit", "Dynamic rate adjustment not supported in standard product"],
    ["REQ-010", "Reporting", "Real-time management dashboards with drill-down capability", "Medium", "Could", "No Fit", "Standard reports are batch-based; no real-time drill-down"],
]
for r, row in enumerate(samples1, 2):
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
style_body(ws1, 2, 6, 7)

add_dv(ws1, ["High", "Medium", "Low"], "D")
add_dv(ws1, ["Must", "Should", "Could", "Won't"], "E")
add_dv(ws1, ["Full Fit", "Partial Fit", "No Fit", "Not Assessed"], "F")

# ========== Sheet 2: Workshop Scoring ==========
ws2 = wb.create_sheet("Workshop Scoring")
headers2 = ["Req ID", "Description", "Fit Rating", "Gap Description", "Resolution Type", "Effort (Days)", "Risk", "Notes", "Parking Lot?"]
widths2 = [12, 45, 14, 45, 18, 12, 10, 40, 12]
style_header(ws2, headers2, widths2)

samples2 = [
    ["REQ-001", "Digital KYC with biometric checks", "Full Fit", "", "Configuration", 5, "Low", "Standard module confirmed by Temenos SME via live demo", "No"],
    ["REQ-002", "PEP and sanctions screening", "Full Fit", "", "Configuration", 3, "Low", "Compliance module demonstrated; lists refreshed daily", "No"],
    ["REQ-003", "Multi-currency with real-time FX", "Partial Fit", "Platform supports multi-currency but FX rates require external provider integration", "Extension", 15, "Medium", "Need FX provider selection before final estimate", "No"],
    ["REQ-008", "Dynamic interest rate adjustment", "No Fit", "Platform supports rate changes via batch but not event-driven rate adjustment based on real-time risk events", "Customisation", 40, "High", "Candidate for CUS-001; score below 5.0 expected", "No"],
    ["REQ-010", "Real-time dashboards with drill-down", "No Fit", "Standard reports are batch-based; drill-down requires custom UI layer over reporting APIs", "Customisation", 35, "Medium", "Candidate for CUS-002; UI-only customisation", "No"],
]
for r, row in enumerate(samples2, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
style_body(ws2, 2, 6, 9)

add_dv(ws2, ["Full Fit", "Partial Fit", "No Fit", "Not Assessed", "Deferred"], "C")
add_dv(ws2, ["Configuration", "Extension", "Customisation", "Process Change", "Workaround", "Deferred"], "E")
add_dv(ws2, ["High", "Medium", "Low"], "G")
add_dv(ws2, ["Yes", "No"], "I")

# ========== Sheet 3: Customisation Scoring ==========
ws3 = wb.create_sheet("Customisation Scoring")
headers3 = ["CUS ID", "Req ID", "Description", "Type",
            "Upgrade Compat (30%)", "Business Crit (25%)", "TCO (20%)",
            "Workaround Feas (15%)", "Impl Risk (10%)",
            "Weighted Score", "Threshold", "Approver", "Status"]
widths3 = [10, 10, 40, 14, 16, 16, 12, 16, 14, 14, 20, 20, 16]
style_header(ws3, headers3, widths3)

samples3 = [
    ["CUS-001", "REQ-008", "Dynamic rate adjustment based on real-time risk events", "Core Logic", 3, 8, 4, 2, 5, None, None, "Steering Committee", "Proposed"],
    ["CUS-002", "REQ-010", "Real-time management dashboard with drill-down analytics", "UI", 7, 6, 7, 3, 6, None, None, "Design Authority", "Under Review"],
]
for r, row in enumerate(samples3, 2):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    # Formulas
    ws3.cell(row=r, column=10, value=f"=(E{r}*0.3)+(F{r}*0.25)+(G{r}*0.2)+(H{r}*0.15)+(I{r}*0.1)")
    ws3.cell(row=r, column=11, value=f'=IF(J{r}>=5,"Design Authority","Steering Committee")')

# Apply formulas to rows 4-50 so users can add more
for r in range(4, 51):
    ws3.cell(row=r, column=10, value=f"=IF(COUNT(E{r}:I{r})=5,(E{r}*0.3)+(F{r}*0.25)+(G{r}*0.2)+(H{r}*0.15)+(I{r}*0.1),\"\")")
    ws3.cell(row=r, column=11, value=f'=IF(ISNUMBER(J{r}),IF(J{r}>=5,"Design Authority","Steering Committee"),"")')

style_body(ws3, 2, 50, 13)

add_dv(ws3, ["Core Logic", "UI", "Integration", "Reporting"], "D")
add_dv(ws3, ["Proposed", "Under Review", "Approved", "Rejected", "Implemented"], "M")

# Conditional formatting on Weighted Score (column J)
red_fill = PatternFill("solid", start_color="F8CBAD")
green_fill = PatternFill("solid", start_color="C6EFCE")
ws3.conditional_formatting.add("J2:J50",
    CellIsRule(operator="lessThan", formula=["5"], fill=red_fill))
ws3.conditional_formatting.add("J2:J50",
    CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=green_fill))

# ========== Sheet 4: Parking Lot ==========
ws4 = wb.create_sheet("Parking Lot")
headers4 = ["Item #", "Req ID", "Description", "Reason", "Owner", "Deadline", "Status"]
widths4 = [8, 12, 45, 18, 20, 14, 14]
style_header(ws4, headers4, widths4)

samples4 = [
    [1, "REQ-012", "Legacy mainframe data sync during dual-run — need schema and volume info", "Missing Info", "Data Migration Lead", "2026-03-22", "Open"],
]
for r, row in enumerate(samples4, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
style_body(ws4, 2, 2, 7)

add_dv(ws4, ["Time", "Missing Info", "Out of Scope"], "D")
add_dv(ws4, ["Open", "Resolved", "Deferred"], "G")

wb.save(OUTPUT)
print(f"Created: {OUTPUT}")
